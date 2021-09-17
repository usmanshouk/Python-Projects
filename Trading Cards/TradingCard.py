#Usman Shoukat
#201537600
import openpyxl
from openpyxl import Workbook
class Card:
    '''
    This is a class which stores the information about name, type, health points, moves and their respective damage, and the shiny status of cards. 
    '''
    def __init__(self, theName, theType, theHP, theMoves, isShiny):
        '''
        This is method to initialize the attributes of Card Class.

        Parameters:
        theName: Name of the card
        theType: Type of the card
        theHP: Health Points of the card
        theMoves: Moves of the card
        isShing: Shiny status of the card

        Return: Nothing
        '''
        self.name = theName
        self.type = theType
        self.healthPoints = theHP
        self.moves = theMoves
        self.shiny = isShiny
    def __str__(self) -> str:
        '''
        This method gives the descrition of given card object.
        '''
        return "This is a card with name " + self.name + " and type " + self.type + " and having health points " + str(self.healthPoints) + ". This card can play " + str(len(self.moves)) + " moves which can make average damage of " + str(self.averageDamage()) + "."
    def averageDamage(self):
        '''
        This method returns the average of damage factors of given card object.

        Parameters: 
        
        self: Object instance

        Return: Average of damages of caused by all moves
        '''
        return sum([i[1] for i in self. moves])/len(self.moves)
class Deck:
    '''
    This is class which stores all the card object instances in a list collection.
    '''
    def __init__(self) -> None:
        '''
        This method initializes the attribute of Deck class.
        '''
        self.deckOfCards = []
    def inputFromFile(self,fileName):
        '''
        This method takes an excel file from user and create the Card instances for each row of excel file by calling the Card Class and stores these instances in the form of a list.

        Parameters:
        
        self: Object instance
        fileName: Excel file

        Return: Nothing
        '''
        #Initialinzing of active_sheet  by active sheet of the given excel file 
        active_sheet = openpyxl.load_workbook(filename = fileName).active
        #iterating through every row of given file and storing each row in the form of tuple in value variable
        for value in active_sheet.iter_rows(min_row = 2, max_row = active_sheet.max_row, min_col = 1, max_col = active_sheet.max_column, values_only = True):
            moves = []
            i = 4
            #iterating through the columns of moves of each row and storing their information in the form of a list
            while i < active_sheet.max_column :
                #Data Cleaning
                #only saving moves which have a move name
                if value[i]:
                    moves.append((value[i],value[i+1]))
                i+=2
            #Data Cleaning
            #Assumption: if Name of the card is not available then data for that tuple is not valid
            if value[0]:
                #Calling of Card class
                card = Card(value[0],value[1],value[2],moves,value[3])
                #adding the card object in the deck
                self.deckOfCards.append(card)
    def __str__(self) -> str:
        '''
        This method gives the descrition of given Deck object.
        '''
        counterOfShinyCards = 0
        #Iterating through all cards to check shiny status of dach card object
        for i in self.deckOfCards:
            if i.shiny == 1:
                counterOfShinyCards += 1
        return "This is a deck with " + str(len(self.deckOfCards)) + " cards. The number of shiny cards in this deck are " + str(counterOfShinyCards) + " and the average damage factor of this deck is " + str(self.getAverageDamage()) + "."
    def addCard(self, theCard):
        '''
        This method adds the asked card object in the collection card objects.

        Parameters:
        
        self: Object instance
        theCard: the card to be added in the collection of cards

        Return: Nothing
        '''
        #Adding the asked card in the deck
        self.deckOfCards.append(theCard)
    def rmCard(self, theCard):
        '''
        This method removes the asked card object from the collection card objects.

        Parameters:
        
        self: Object instance
        theCard: the card to be removes from the collection of cards

        Return: Nothing
        '''
        #Removing the asked card from the deck
        self.deckOfCards.remove(theCard)
    def getMostPowerful(self):
        '''
        This method returns the most powerful card in the deck.

        Parameters:
        
        self: Object instance

        Return: Most powerful card object
        '''
        damagePower = 0
        #Iterating through every card object
        for i in range(0, len(self.deckOfCards)):
            #storing average damage factor of each card in variable a
            a = self.deckOfCards[i].averageDamage()
            #Comparing the average damage factor of current card object with previous card with highest damage factor
            if a > damagePower:
                #if the damage factor of current cad is higher then storing it as highest damage power untill the current iteration
                damagePower = a
                #saving of index of powerful card object until the current iteration in variable index
                index = i
        return self.deckOfCards[index]

    def getAverageDamage(self):
        '''
        This method returns the average of damage factor of the asked deck.

        Parameters:
        
        self: Object instance

        Return: Average of damage factor of given deck
        '''
        #returing average damage factor of deck upto 1 decimal places
        return round(sum([i.averageDamage() for i in self.deckOfCards])/len(self.deckOfCards),1)
    def viewAllCards(self):
        '''
        This method prints this information of all the cards in the deck.

        Parameters:
        
        self: Object instance

        Return: Nothing
        '''
        print("\n************ All Cards ************\n")
        for i in self.deckOfCards:
            print("   ",i.name)
            print("Type: ",i.type) 
            print("HP: ",i.healthPoints)
            print("Moves:")
            for j in i.moves:
                print("     Move: ",j[0],"=>  Damage =",j[1])
            print("Shiny: ", i.shiny)
            print("\t--------------------------- ")
            print()
        print("\n\t   ********************* ")
    def viewAllShinyCards(self):
        '''
        This method prints this information of all the shiny cards in the deck.

        Parameters:
        
        self: Object instance

        Return: Nothing
        '''
        print("\n*********** Shiny Cards ***********\n")
        for i in self.deckOfCards:
            if i.shiny == True:
                print("   ",i.name)
                print("Type: ",i.type) 
                print("HP: ",i.healthPoints)
                print("Moves:")
                for j in i.moves:
                    print("     Move: ",j[0],"=>  Damage =",j[1])
                print("Shiny: ", i.shiny)
                print("\t--------------------------- ")
                print()
        print("\n\t   ********************* ")
    def viewAllByType(self, theType):
        '''
        This method prints this information of all the cards with specific type in the deck.

        Parameters:
        
        self: Object instance
        theType: type for which cards to be displayed

        Return: Nothing
        '''
        print("\n************* Cards of type", theType, "*************\n")
        for i in self.deckOfCards:
            if i.type == theType:
                print("   ",i.name)
                print("Type: ",i.type) 
                print("HP: ",i.healthPoints)
                print("Moves:")
                for j in i.moves:
                    print("     Move: ",j[0],"=>  Damage =",j[1])
                print("Shiny: ", i.shiny)
                print("\t--------------------------- ")
                print()
        print("\n\t   ********************* ")
    def getCards(self):
        '''
        This method returns the collection of all the card objects.

        Parameters:
        
        self: Object instance

        Return: collection of all the card objects
        '''
        return self.deckOfCards
    def saveToFile(self, fileName):
        '''
        This method saves all the cards in the deck in the given excel file.

        Parameters:
        
        self: Object instance
        fileName: the file in which cards to be stored

        Return: Nothing
        '''
        workbook = Workbook()
        sheet = workbook.active
        attributes = ['Name', 'Type', 'HP', 'Shiny', 'Move Name 1', 'Damage 1', 'Move Name 2', 'Damage 2', 'Move Name 3', 'Damage 3', 'Move Name 4', 'Damage 4', 'Move Name 5', 'Damage 5']
        #iteraing through every element of attributes and storing them in each column of first row
        for i in range(1, len(attributes)+1):
            sheet.cell(row=1,column=i).value = attributes[i-1]
        #iterating through every card object in the deck storing it in each card in each row 
        for i in range(2, len(self.deckOfCards)+2):
            sheet.cell(row = i,column = 1).value = self.deckOfCards[i-2].name
            sheet.cell(row = i,column = 2).value = self.deckOfCards[i-2].type
            sheet.cell(row = i,column = 3).value = self.deckOfCards[i-2].healthPoints
            sheet.cell(row = i,column = 4).value = self.deckOfCards[i-2].shiny
            j = 0
            l = 5
            #iterating through the list of moves and storing the moves in different columns of given row 
            while j < len(self.deckOfCards[i-2].moves):
                for k in range(0,2):
                    sheet.cell(row = i,column = l).value = self.deckOfCards[i-2].moves[j][k]
                    l +=1
                j += 1
        #saving to given file
        workbook.save(filename=fileName)
def main():
    pass
if __name__ == "__main__":
    main()